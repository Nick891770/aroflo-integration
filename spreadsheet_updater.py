"""
Spreadsheet Updater for AroFlo Scorecard

Updates a scorecard spreadsheet with monthly metrics from AroFlo API.
Customise SPREADSHEET_PATH, MONTH_COLUMN_MAP, and METRIC_ROWS to match
your own spreadsheet layout.
"""

import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

from data_extractor import MonthlyMetrics


# Path to your scorecard spreadsheet - configure via environment or edit directly
SPREADSHEET_PATH = os.getenv(
    "SCORECARD_PATH",
    os.path.join(os.path.dirname(__file__), "scorecard.xlsx"),
)

# Month column mapping based on spreadsheet structure
# Each month spans 2 columns: Target and Actual
# Adjust column numbers to match your spreadsheet layout
MONTH_COLUMN_MAP = {
    11: {"name": "NOVEMBER", "target": 4, "actual": 5},   # November
    12: {"name": "DECEMBER", "target": 6, "actual": 7},   # December
    1: {"name": "JANUARY", "target": 8, "actual": 9},     # January
    2: {"name": "FEBRUARY", "target": 10, "actual": 11},  # February
    3: {"name": "MARCH", "target": 12, "actual": 13},     # March
    4: {"name": "APRIL", "target": 14, "actual": 15},     # April
    5: {"name": "MAY", "target": 16, "actual": 17},       # May
    6: {"name": "JUNE", "target": 18, "actual": 19},      # June
    7: {"name": "JULY", "target": 20, "actual": 21},      # July
    8: {"name": "AUGUST", "target": 22, "actual": 23},    # August
    9: {"name": "SEPTEMBER", "target": 24, "actual": 25}, # September
    10: {"name": "OCTOBER", "target": 26, "actual": 27},  # October
}

# Row numbers for metrics - adjust to match your spreadsheet layout
# LAGGING/PAST section - update Actual column only
METRIC_ROWS = {
    # LAGGING/PAST metrics (rows that need data, not formulas)
    "revenue": 8,               # Revenue/Sales Income
    "gross_profit_dollars": 10, # Gross Profit $
    "net_profit_dollars": 12,   # Net Profit $
    # Row 14: Gross Profit % - FORMULA (skip)
    # Row 16: Net Profit % - FORMULA (skip)
    "completed_jobs": 18,       # Number of completed jobs
    "average_job_value": 20,    # Average job value

    # LEADING/PREDICTIVE metrics
    "primary_client_jobs": 24,      # Primary Client - # jobs
    "primary_client_value": 26,     # Primary Client - $ value
    "other_client_jobs": 28,        # Other Clients - # jobs
    "other_client_value": 30,       # Other Clients - $ value
    # Row 32: Jobs Total - FORMULA (skip)
    # Row 34: % Of Sales from primary client - FORMULA (skip)
    # Row 36: % Of Sales from other clients - FORMULA (skip)
}


class SpreadsheetUpdater:
    """Updates a scorecard spreadsheet with monthly metrics."""

    def __init__(self, spreadsheet_path: Optional[str] = None):
        """
        Initialize the spreadsheet updater.

        Args:
            spreadsheet_path: Path to the Excel file (uses default if not provided)
        """
        self.spreadsheet_path = Path(spreadsheet_path or SPREADSHEET_PATH)

    def _backup_spreadsheet(self) -> Path:
        """
        Create a backup of the spreadsheet before modifying.

        Returns:
            Path to the backup file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self.spreadsheet_path.with_suffix(f".backup_{timestamp}.xlsx")
        shutil.copy2(self.spreadsheet_path, backup_path)
        print(f"Created backup: {backup_path}")
        return backup_path

    def _get_actual_column(self, month: int) -> int:
        """
        Get the Actual column number for a given month.

        Args:
            month: Month number (1-12)

        Returns:
            Column number for the Actual column

        Raises:
            ValueError: If month is invalid
        """
        if month not in MONTH_COLUMN_MAP:
            raise ValueError(f"Invalid month: {month}. Must be 1-12.")
        return MONTH_COLUMN_MAP[month]["actual"]

    def update_spreadsheet(
        self,
        metrics: MonthlyMetrics,
        month: int,
        create_backup: bool = True,
    ) -> bool:
        """
        Update the spreadsheet with monthly metrics.

        Args:
            metrics: MonthlyMetrics object with calculated values
            month: Month number (1-12)
            create_backup: Whether to create a backup before updating

        Returns:
            True if update was successful, False otherwise
        """
        if not self.spreadsheet_path.exists():
            print(f"Error: Spreadsheet not found at {self.spreadsheet_path}")
            return False

        try:
            # Create backup
            if create_backup:
                self._backup_spreadsheet()

            # Load workbook
            wb = openpyxl.load_workbook(self.spreadsheet_path)
            sheet = wb.active

            # Get the actual column for this month
            actual_col = self._get_actual_column(month)
            month_name = MONTH_COLUMN_MAP[month]["name"]

            print(f"\nUpdating {month_name} (column {actual_col})...")

            # Update metrics
            updates = [
                ("Revenue/Sales Income", METRIC_ROWS["revenue"], metrics.revenue),
                ("Gross Profit $", METRIC_ROWS["gross_profit_dollars"], metrics.gross_profit_dollars),
                ("Net Profit $", METRIC_ROWS["net_profit_dollars"], metrics.net_profit_dollars),
                ("Number of completed jobs", METRIC_ROWS["completed_jobs"], metrics.completed_jobs),
                ("Average job value", METRIC_ROWS["average_job_value"], metrics.average_job_value),
                ("Primary Client - # jobs", METRIC_ROWS["primary_client_jobs"], metrics.primary_client_jobs),
                ("Primary Client - $ value", METRIC_ROWS["primary_client_value"], metrics.primary_client_value),
                ("Other Clients - # jobs", METRIC_ROWS["other_client_jobs"], metrics.other_client_jobs),
                ("Other Clients - $ value", METRIC_ROWS["other_client_value"], metrics.other_client_value),
            ]

            for name, row, value in updates:
                cell = sheet.cell(row=row, column=actual_col)
                cell.value = value
                print(f"  {name}: {value}")

            # Save the workbook
            wb.save(self.spreadsheet_path)
            wb.close()

            print(f"\nSpreadsheet updated successfully!")
            return True

        except Exception as e:
            print(f"Error updating spreadsheet: {e}")
            return False

    def get_current_values(self, month: int) -> dict:
        """
        Get current values from the spreadsheet for a given month.

        Args:
            month: Month number (1-12)

        Returns:
            Dictionary of current metric values
        """
        if not self.spreadsheet_path.exists():
            return {}

        try:
            wb = openpyxl.load_workbook(self.spreadsheet_path, data_only=True)
            sheet = wb.active

            actual_col = self._get_actual_column(month)

            values = {}
            for name, row in METRIC_ROWS.items():
                cell = sheet.cell(row=row, column=actual_col)
                values[name] = cell.value

            wb.close()
            return values

        except Exception as e:
            print(f"Error reading spreadsheet: {e}")
            return {}


def update_spreadsheet(metrics: MonthlyMetrics, month: int) -> bool:
    """
    Convenience function to update the spreadsheet.

    Args:
        metrics: MonthlyMetrics object with calculated values
        month: Month number (1-12)

    Returns:
        True if update was successful, False otherwise
    """
    updater = SpreadsheetUpdater()
    return updater.update_spreadsheet(metrics, month)
